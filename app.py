from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort
from flask_sqlalchemy import SQLAlchemy
import os
import pandas as pd
import io
import csv
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session
from datetime import datetime
import zipfile
import openpyxl
import math

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db.sqlite3'
app.config['UPLOAD_FOLDER'] = 'uploads/questions'

db = SQLAlchemy(app)

# Models
class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    text = db.Column(db.String, nullable=False)
    option_a = db.Column(db.String, nullable=False)
    option_b = db.Column(db.String, nullable=False)
    option_c = db.Column(db.String, nullable=False)
    option_d = db.Column(db.String, nullable=False)
    correct_option = db.Column(db.String(1), nullable=False)  # 'A', 'B', 'C', or 'D'

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String, nullable=False)
    email = db.Column(db.String, nullable=True)

class Response(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('student.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    selected_option = db.Column(db.String(1), nullable=False)
    is_correct = db.Column(db.Boolean, nullable=False)

    student = db.relationship('Student', backref=db.backref('responses', lazy=True))
    question = db.relationship('Question', backref=db.backref('responses', lazy=True))

# User model for authentication
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(10), nullable=False)  # 'admin' or 'student'
    name = db.Column(db.String(100), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Admin: Upload questions and model answers
@app.route('/admin/upload', methods=['GET', 'POST'])
def admin_upload():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('No file selected', 'error')
            return redirect(request.url)
        try:
            filename = file.filename
            if filename.endswith('.csv'):
                df = pd.read_csv(file)
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                df = pd.read_excel(file)
            else:
                flash('Unsupported file type. Please upload a CSV or Excel file.', 'error')
                return redirect(request.url)
            required_cols = ['Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Option']
            if not all(col in df.columns for col in required_cols):
                flash('Missing required columns in file.', 'error')
                return redirect(request.url)
            # Clear existing questions
            Question.query.delete()
            db.session.commit()
            # Add questions
            for _, row in df.iterrows():
                q = Question(
                    text=row['Question'],
                    option_a=row['Option A'],
                    option_b=row['Option B'],
                    option_c=row['Option C'],
                    option_d=row['Option D'],
                    correct_option=row['Correct Option'].strip().upper()
                )
                db.session.add(q)
            db.session.commit()
            flash('Questions uploaded successfully!', 'success')
            return redirect(request.url)
        except Exception as e:
            flash(f'Error processing file: {e}', 'error')
            return redirect(request.url)
    return render_template('admin_upload.html')

# Student: Take the test
@app.route('/test', methods=['GET', 'POST'])
def student_test():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email') or None
        if not name:
            flash('Name is required.', 'error')
            return redirect(request.url)
        # Check if student already exists
        student = Student.query.filter_by(email=email).first() if email else None
        if not student:
            student = Student(name=name, email=email)
            db.session.add(student)
            db.session.commit()
        # Prevent multiple attempts (optional)
        if Response.query.filter_by(student_id=student.id).first():
            flash('You have already taken the test.', 'error')
            return redirect(request.url)
        questions = Question.query.all()
        score = 0
        for q in questions:
            selected = request.form.get(f'q_{q.id}')
            is_correct = selected == q.correct_option
            if is_correct:
                score += 1
            response = Response(
                student_id=student.id,
                question_id=q.id,
                selected_option=selected,
                is_correct=is_correct
            )
            db.session.add(response)
        db.session.commit()
        return render_template('student_test.html', submitted=True, score=score, total=len(questions))
    else:
        questions = Question.query.all()
        return render_template('student_test.html', questions=questions, submitted=False)

# Admin: View results
@app.route('/admin/results')
def view_results():
    students = Student.query.all()
    results = []
    for student in students:
        responses = Response.query.filter_by(student_id=student.id).all()
        score = sum(1 for r in responses if r.is_correct)
        results.append({
            'name': student.name,
            'email': student.email,
            'score': score,
            'total': len(responses)
        })
    return render_template('results.html', results=results)

@app.route('/admin/results/download')
def download_results():
    students = Student.query.all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Email', 'Score', 'Total Answered'])
    for student in students:
        responses = Response.query.filter_by(student_id=student.id).all()
        score = sum(1 for r in responses if r.is_correct)
        writer.writerow([student.name, student.email or '', score, len(responses)])
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name='results.csv'
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user)
            flash('Logged in successfully!', 'success')
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        else:
            flash('Invalid email or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        if User.query.filter_by(email=email).first():
            flash('Email already registered.', 'danger')
            return redirect(url_for('register'))
        user = User(email=email, name=name, role='student')
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Access denied.', 'danger')
        return redirect(url_for('login'))
    return render_template('admin_dashboard.html')

@app.route('/student/dashboard')
@login_required
def student_dashboard():
    if current_user.role != 'student':
        flash('Access denied.', 'danger')
        return redirect(url_for('login'))
    return render_template('student_dashboard.html')

@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        abort(403)
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
def admin_add_user():
    if current_user.role != 'admin':
        abort(403)
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        if User.query.filter_by(email=email).first():
            flash('Email already registered.', 'danger')
            return redirect(url_for('admin_add_user'))
        user = User(email=email, name=name, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('User added successfully!', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin_add_user.html')

@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_user(user_id):
    if current_user.role != 'admin':
        abort(403)
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.name = request.form.get('name')
        user.email = request.form.get('email')
        role = request.form.get('role')
        user.role = role
        password = request.form.get('password')
        if password:
            user.set_password(password)
        db.session.commit()
        flash('User updated successfully!', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin_edit_user.html', user=user)

@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    if current_user.role != 'admin':
        abort(403)
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot delete yourself.', 'danger')
        return redirect(url_for('admin_users'))
    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully!', 'success')
    return redirect(url_for('admin_users'))

class Test(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False)
    test_id = db.Column(db.String(50), unique=True, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    filename = db.Column(db.String(200), nullable=False)  # uploaded ZIP filename
    admin = db.relationship('User', backref=db.backref('tests', lazy=True))

@app.route('/admin/tests')
@login_required
def admin_tests():
    if current_user.role != 'admin':
        abort(403)
    tests = Test.query.order_by(Test.created_at.desc()).all()
    return render_template('admin_tests.html', tests=tests)

@app.route('/admin/tests/create', methods=['GET', 'POST'])
@login_required
def admin_create_test():
    if current_user.role != 'admin':
        abort(403)
    if request.method == 'POST':
        subject = request.form.get('subject')
        test_id = request.form.get('test_id')
        file = request.files.get('file')
        if not subject or not test_id or not file:
            flash('All fields are required.', 'danger')
            return redirect(url_for('admin_create_test'))
        if not file.filename.endswith('.zip'):
            flash('Please upload a ZIP file.', 'danger')
            return redirect(url_for('admin_create_test'))
        # Save ZIP file
        zip_filename = f"{test_id}_{file.filename}"
        zip_path = os.path.join(app.config['UPLOAD_FOLDER'], zip_filename)
        file.save(zip_path)
        # Optionally extract and parse here
        test = Test(subject=subject, test_id=test_id, created_by=current_user.id, filename=zip_filename)
        db.session.add(test)
        db.session.commit()
        flash('Test created and question paper uploaded!', 'success')
        return redirect(url_for('admin_tests'))
    return render_template('admin_create_test.html')

@app.route('/admin/tests/delete/<int:test_id>', methods=['POST'])
@login_required
def admin_delete_test(test_id):
    if current_user.role != 'admin':
        abort(403)
    test = Test.query.get_or_404(test_id)
    # Optionally delete the uploaded file
    try:
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], test.filename))
    except Exception:
        pass
    db.session.delete(test)
    db.session.commit()
    flash('Test deleted.', 'success')
    return redirect(url_for('admin_tests'))

@app.route('/student/tests')
@login_required
def student_tests():
    if current_user.role != 'student':
        abort(403)
    tests = Test.query.order_by(Test.created_at.desc()).all()
    return render_template('student_tests.html', tests=tests)

@app.route('/student/test/<test_id>/start', methods=['GET'])
@login_required
def student_start_test(test_id):
    if current_user.role != 'student':
        abort(403)
    test = Test.query.filter_by(test_id=test_id).first_or_404()
    # Parse the Excel file from the ZIP (for demo, just list questions)
    zip_path = os.path.join(app.config['UPLOAD_FOLDER'], test.filename)
    questions = []
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        excel_file = None
        for name in zip_ref.namelist():
            if name.endswith('.xlsx') or name.endswith('.xls'):
                excel_file = name
                break
        if excel_file:
            with zip_ref.open(excel_file) as f:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    q = dict(zip(headers, row))
                    questions.append(q)
    return render_template('student_test_start.html', test=test, questions=questions)

@app.route('/student/test/<test_id>/jee', methods=['GET', 'POST'])
@login_required
def student_jee_test(test_id):
    if current_user.role != 'student':
        abort(403)
    test = Test.query.filter_by(test_id=test_id).first_or_404()
    zip_path = os.path.join(app.config['UPLOAD_FOLDER'], test.filename)
    # Parse questions from Excel in ZIP
    questions = []
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        excel_file = None
        for name in zip_ref.namelist():
            if name.endswith('.xlsx') or name.endswith('.xls'):
                excel_file = name
                break
        if excel_file:
            with zip_ref.open(excel_file) as f:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    q = dict(zip(headers, row))
                    questions.append(q)
    num_questions = len(questions)
    # State: current question index, answers, review status
    if 'jee_state' not in session or session.get('jee_test_id') != test_id:
        session['jee_state'] = {
            'current': 0,
            'answers': [None] * num_questions,
            'review': [False] * num_questions,
            'visited': [False] * num_questions,
            'start_time': math.floor(session.get('jee_start_time', 0) or 0),
        }
        session['jee_test_id'] = test_id
        session['jee_start_time'] = math.floor(session.get('jee_start_time', 0) or 0)
    state = session['jee_state']
    if request.method == 'POST':
        action = request.form.get('action')
        q_idx = int(request.form.get('q_idx'))
        selected = request.form.get('selected')
        # Save answer if selected
        if selected:
            state['answers'][q_idx] = selected
        state['visited'][q_idx] = True
        # Mark for review
        if action == 'mark_review':
            state['review'][q_idx] = True
        elif action == 'save_next':
            state['review'][q_idx] = False
            state['current'] = min(q_idx + 1, num_questions - 1)
        elif action == 'clear':
            state['answers'][q_idx] = None
            state['review'][q_idx] = False
        elif action == 'goto':
            goto_idx = int(request.form.get('goto_idx'))
            state['current'] = goto_idx
        elif action == 'submit':
            # TODO: Save responses to DB, calculate score, clear session
            return redirect(url_for('student_dashboard'))
        session['jee_state'] = state
    q_idx = state['current']
    return render_template('student_jee_test.html', test=test, questions=questions, q_idx=q_idx, state=state, enumerate=enumerate, len=len)

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True) 